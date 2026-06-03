"""Dump every EN 16931 / Factur-X 1.08 business rule into a JSON sidecar.

Reads three sources inside the Factur-X workbook:

* ``Business Rules`` sheet, **right half** (cols 13-25) — the
  ``BR-1..65``, ``BR-CO-*``, ``BR-CL-*``, ``BR-DEC-*``,
  ``BR-FXEXT-*`` and Italian ``BR-B-*`` "integrity constraints and
  conditions" family. Each row carries the EN + FR description, the
  target context, the BT/BG the rule references, and a per-profile
  X-mark matrix.
* ``Business Rules`` sheet, **left half** (cols 3-11) — the
  per-VAT-category families (``BR-S/Z/E/AE/G/IC/AF/AG/O-1..14``).
  Same shape minus the target / business-term columns.
* ``Business Rules HYBRID`` sheet — the Factur-X hybrid-PDF
  ``BR-HYBRID-*`` rules, carried trilingually (EN / FR / DE) with
  applicability, error level and comment columns instead of a
  profile matrix.

Source:
    By default the vendored workbook at
    ``docs/spec/1_FACTUR-X 1.08 - 2025 12 04 - EN FR - VF.xlsx``;
    override via the positional CLI argument.

Output:
    ``tools/business_rules.json`` by default; override with ``--out``.
    The file is a JSON object keyed by BR id.

Run:
    uv run python tools/extract_business_rules.py
    uv run python tools/extract_business_rules.py path/to/workbook.xlsx
    uv run python tools/extract_business_rules.py --out rules.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

DEFAULT_OUT = ROOT / "tools/business_rules.json"

# ---------------------------------------------------------------------------
# "Business Rules" sheet — header row 4, data starts row 5.
# ---------------------------------------------------------------------------
#
# Left half (per-VAT-category):    | right half (general / arithmetic):
#   col 3  ID                      |   col 13 ID
#   col 4  Description EN          |   col 14 Description EN
#   col 5  Description FR          |   col 15 Target / context EN
#   col 7..11  profile X-marks     |   col 16 Business term / group EN
#                                  |   col 17 Description FR
#                                  |   col 18 Cible / contexte FR
#                                  |   col 19 Terme métier / groupe FR
#                                  |   col 21..25 profile X-marks

SHEET_GENERAL = "Business Rules"
SHEET_HYBRID = "Business Rules HYBRID"

VAT_COLUMNS: dict[str, int] = {"id": 3, "description": 4, "description_fr": 5}
VAT_PROFILE_COLUMNS: dict[str, int] = {
    "MINIMUM": 7,
    "BASIC_WL": 8,
    "BASIC": 9,
    "EN16931": 10,
    "EXTENDED": 11,
}

GENERAL_COLUMNS: dict[str, int] = {
    "id": 13,
    "description": 14,
    "target": 15,
    "business_term": 16,
    "description_fr": 17,
    "target_fr": 18,
    "business_term_fr": 19,
}
GENERAL_PROFILE_COLUMNS: dict[str, int] = {
    "MINIMUM": 21,
    "BASIC_WL": 22,
    "BASIC": 23,
    "EN16931": 24,
    "EXTENDED": 25,
}

# ---------------------------------------------------------------------------
# "Business Rules HYBRID" sheet — header row 4, data starts row 5.
# Three trilingual blocks side by side: EN (cols 3-7), FR (9-13), DE (15-19).
# Each block has: BR id, Text, Applicable, Error level, Comment.
# HYBRID rules don't carry a per-profile matrix (they apply to every
# Factur-X hybrid document irrespective of profile).
# ---------------------------------------------------------------------------

HYBRID_EN: dict[str, int] = {
    "id": 3,
    "description": 4,
    "applicable": 5,
    "error_level": 6,
    "comment": 7,
}
HYBRID_FR: dict[str, int] = {
    "id_fr": 9,
    "description_fr": 10,
    "applicable_fr": 11,
    "error_level_fr": 12,
    "comment_fr": 13,
}
HYBRID_DE: dict[str, int] = {
    "id_de": 15,
    "description_de": 16,
    "applicable_de": 17,
    "error_level_de": 18,
    "comment_de": 19,
}

ID_RE = re.compile(r"^BR(?:-[A-Z0-9]+)+$")


def _norm(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text == "-":
        return None
    paragraphs = [re.sub(r"\s+", " ", p).strip() for p in text.split("\n")]
    paragraphs = [p for p in paragraphs if p]
    return "\n".join(paragraphs) or None


def _cell(row: list[Any], col: int) -> Any:
    return row[col - 1] if col - 1 < len(row) else None


def _row_to_dict(row: list[Any], columns: dict[str, int]) -> dict[str, Any]:
    return {key: _norm(_cell(row, col)) for key, col in columns.items()}


def _profiles_present(row: list[Any], columns: dict[str, int]) -> list[str]:
    return [
        name
        for name, col in columns.items()
        if _cell(row, col) is not None and str(_cell(row, col)).strip().upper() == "X"
    ]


def _family(rule_id: str) -> str:
    """Group rules so consumers can filter by family.

    ``BR-S-1`` → ``BR-S``; ``BR-CO-25`` → ``BR-CO``;
    ``BR-FXEXT-CO-10`` → ``BR-FXEXT-CO``; ``BR-1`` → ``BR``.
    Numeric trailing segments are stripped one at a time.
    """
    parts = rule_id.split("-")
    while parts and parts[-1].isdigit():
        parts.pop()
    return "-".join(parts) or rule_id


def _extract_general(ws: Any, terms: dict[str, dict[str, Any]]) -> None:
    """Walk the right half of the ``Business Rules`` sheet."""
    for row_number, raw_row in enumerate(
        ws.iter_rows(min_row=5, values_only=True), start=5
    ):
        row = list(raw_row)
        rule_id = _norm(_cell(row, GENERAL_COLUMNS["id"]))
        if not rule_id or not ID_RE.match(rule_id):
            continue

        entry = _row_to_dict(row, GENERAL_COLUMNS)
        entry["id"] = rule_id
        entry["family"] = _family(rule_id)
        entry["profiles"] = _profiles_present(row, GENERAL_PROFILE_COLUMNS)
        entry["source_section"] = "general"
        entry["source_sheet"] = SHEET_GENERAL
        entry["source_row"] = row_number
        terms[rule_id] = entry


def _extract_vat(ws: Any, terms: dict[str, dict[str, Any]]) -> None:
    """Walk the left half of the ``Business Rules`` sheet."""
    for row_number, raw_row in enumerate(
        ws.iter_rows(min_row=5, values_only=True), start=5
    ):
        row = list(raw_row)
        rule_id = _norm(_cell(row, VAT_COLUMNS["id"]))
        if not rule_id or not ID_RE.match(rule_id):
            continue

        entry = _row_to_dict(row, VAT_COLUMNS)
        entry["id"] = rule_id
        entry["family"] = _family(rule_id)
        entry["profiles"] = _profiles_present(row, VAT_PROFILE_COLUMNS)
        entry["source_section"] = "vat"
        entry["source_sheet"] = SHEET_GENERAL
        entry["source_row"] = row_number
        terms[rule_id] = entry


def _extract_hybrid(ws: Any, terms: dict[str, dict[str, Any]]) -> None:
    """Walk the ``Business Rules HYBRID`` sheet (trilingual)."""
    for row_number, raw_row in enumerate(
        ws.iter_rows(min_row=5, values_only=True), start=5
    ):
        row = list(raw_row)
        rule_id = _norm(_cell(row, HYBRID_EN["id"]))
        if not rule_id or not ID_RE.match(rule_id):
            continue

        entry = _row_to_dict(row, HYBRID_EN)
        entry.update(_row_to_dict(row, HYBRID_FR))
        entry.update(_row_to_dict(row, HYBRID_DE))
        entry["id"] = rule_id
        entry["family"] = _family(rule_id)
        entry["source_section"] = "hybrid"
        entry["source_sheet"] = SHEET_HYBRID
        entry["source_row"] = row_number
        terms[rule_id] = entry


def extract(xlsx: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    terms: dict[str, dict[str, Any]] = {}

    if SHEET_GENERAL in wb.sheetnames:
        ws = wb[SHEET_GENERAL]
        _extract_general(ws, terms)
        _extract_vat(ws, terms)
    else:
        print(  # noqa: T201
            f"warning: sheet {SHEET_GENERAL!r} not found, skipping", file=sys.stderr
        )

    if SHEET_HYBRID in wb.sheetnames:
        _extract_hybrid(wb[SHEET_HYBRID], terms)
    else:
        print(  # noqa: T201
            f"warning: sheet {SHEET_HYBRID!r} not found, skipping", file=sys.stderr
        )

    return terms


def _build_parser() -> argparse.ArgumentParser:
    assert __doc__
    parser = argparse.ArgumentParser(
        prog="extract_business_rules", description=__doc__.split("\n\n", maxsplit=1)[0]
    )
    _ = parser.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        help="Path to the Factur-X workbook (``1_FACTUR-X 1.08 - ... - VF.xlsx``). ",
    )
    _ = parser.add_argument(
        "-o",
        "--out",
        default=DEFAULT_OUT,
        type=Path,
        help=f"Output JSON path. Defaults to {DEFAULT_OUT.relative_to(ROOT)}.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    xlsx: Path | None = args.xlsx
    out: Path = args.out

    if xlsx is None:
        print("error: you have to provide a XLSX file as input", file=sys.stderr)  # noqa: T201
        return 2

    if not xlsx.is_file():
        print(f"error: cannot read XLSX at {xlsx}", file=sys.stderr)  # noqa: T201
        return 2

    rules = extract(xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    _ = out.write_text(json.dumps(rules, indent=2, sort_keys=True) + "\n")

    from collections import Counter

    families = Counter(r["family"] for r in rules.values())
    rel = (
        out.relative_to(ROOT) if out.is_absolute() and out.is_relative_to(ROOT) else out
    )
    family_summary = ", ".join(
        f"{name}={count}" for name, count in sorted(families.items())
    )
    print(  # noqa: T201
        f"wrote {len(rules)} rules to {rel}\n  families: {family_summary}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
