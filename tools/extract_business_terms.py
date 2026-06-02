"""Dump every EN 16931 / Factur-X 1.08 business term into JSON sidecars.

Walks each per-profile sheet in the Factur-X workbook
(``Factur-X CII D22B {MINIMUM,BASIC WL,BASIC,EN16931,EXTENDED}``) and
emits one entry per BT/BG id, carrying every column the workbook
provides: scalar metadata (block, data type, xpath, cardinality,
business rules) in both English and French, the per-profile X-mark
matrix, the parent / child / depth columns, and a per-occurrence
list recording every (sheet, row) the term appeared on so consumers
can disambiguate when a term shows up multiple times (header vs line
context, for instance).

Source:
    By default the vendored workbook at
    ``docs/spec/1_FACTUR-X 1.08 - 2025 12 04 - EN FR - VF.xlsx``;
    override via the positional CLI argument.

Outputs:
    * ``tools/business_terms.json`` — flat JSON object keyed by BT/BG id.
    * ``tools/business_terms_tree.json`` — same data, restructured as
      a tree keyed by the xpath segment (single-step keys like
      ``/rsm:ExchangedDocument``, ``/ram:ID``, ``/@schemeID``). Each
      node carries its ``id`` (the BT/BG number that lives at that
      xpath, ``null`` for pure structural containers), ``name``,
      ``profiles`` (sheets the term appears on), and ``children``.

    Both output paths can be overridden with ``--out`` / ``--out-tree``.

Run:
    uv run python tools/extract_business_terms.py
    uv run python tools/extract_business_terms.py path/to/workbook.xlsx
    uv run python tools/extract_business_terms.py --out terms.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

DEFAULT_XLSX = ROOT / "docs/spec/1_FACTUR-X 1.08 - 2025 12 04 - EN FR - VF.xlsx"
DEFAULT_OUT = ROOT / "tools/business_terms.json"
DEFAULT_OUT_TREE = ROOT / "tools/business_terms_tree.json"

PROFILE_SHEETS: list[str] = [
    "Factur-X CII D22B MINIMUM",
    "Factur-X CII D22B BASIC WL",
    "Factur-X CII D22B BASIC",
    "Factur-X CII D22B EN16931",
    "Factur-X CII D22B EXTENDED",
]

# 1-based column indices, identical across every profile sheet.
# Header row = row 4; data starts at row 5.
EN_COLUMNS: dict[str, int] = {
    "is_bt": 1,  # "BT (YES) or not (NO)"
    "change_1_0_07": 2,
    "change_1_07_3": 3,
    "change_1_08": 4,
    "block": 5,  # "CDE BLOC CII"
    "id": 6,
    "id_ctc_fr_reform": 7,
    "xsd_level": 8,
    "semantic_cardinality": 9,
    "name": 10,  # "Business Term"
    "description": 11,
    "usage_note": 12,
    "cius_note": 13,
    "business_rules": 14,
    "data_type": 15,
    "xml_cardinality": 16,
    "ext_profiles_cardinality": 17,
    "xpath_raw": 18,
    "xpath": 19,
    "dt": 20,
    "type": 21,
    "cii_cardinality": 22,
    "match": 23,
    "rules": 24,
    "factur_x_lowest": 34,
    "profile_subset_extended": 35,
    "parent": 37,
    "child": 38,
    "nb_parents_above": 39,
}

# French side of the workbook — same column shape, offset by +35.
FR_COLUMNS: dict[str, int] = {
    "block_fr": 41,
    "semantic_cardinality_fr": 45,
    "name_fr": 46,
    "description_fr": 47,
    "usage_note_fr": 48,
    "cius_note_fr": 49,
    "business_rules_fr": 50,
}

# Per-profile X-mark matrix, identical across sheets.
PROFILE_COLUMNS: dict[str, int] = {
    "MINIMUM": 26,
    "BASIC_WL": 27,
    "BASIC": 28,
    "EN16931": 29,
    "EXTENDED-CTC-FR": 30,
    "EXTENDED-B2B-FR": 31,
    "EXTENDED": 32,
}

# Per-sheet "this row appeared in" tag (mirrors PROFILE_COLUMNS keys but
# keyed by the workbook sheet name so consumers can join back to a sheet).
SHEET_TAG: dict[str, str] = {
    "Factur-X CII D22B MINIMUM": "MINIMUM",
    "Factur-X CII D22B BASIC WL": "BASIC_WL",
    "Factur-X CII D22B BASIC": "BASIC",
    "Factur-X CII D22B EN16931": "EN16931",
    "Factur-X CII D22B EXTENDED": "EXTENDED",
}

# Per-occurrence columns — kept on every (sheet, row) entry under
# ``occurrences`` since they vary per occurrence even when the term id
# is the same.
OCCURRENCE_COLUMNS: tuple[str, ...] = (
    "block",
    "xpath",
    "xpath_raw",
    "xsd_level",
    "semantic_cardinality",
    "xml_cardinality",
    "ext_profiles_cardinality",
    "cii_cardinality",
    "dt",
    "type",
    "match",
    "rules",
    "parent",
    "child",
    "nb_parents_above",
)

ID_RE = re.compile(r"^B[GT]-(X-)?\d+(-\d+)?(-\d+)?$")


def _norm(value: Any) -> Any:
    """Strip whitespace; collapse internal whitespace inside each
    paragraph but preserve newline-separated paragraphs.

    Returns ``None`` for empty / placeholder cells. Non-string values
    (numbers, booleans) pass through unchanged.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text == "-":
        return None
    paragraphs = [re.sub(r"\s+", " ", p).strip() for p in text.split("\n")]
    paragraphs = [p for p in paragraphs if p]
    return "\n".join(paragraphs) or None


def _profiles_present(row: list[Any]) -> list[str]:
    present: list[str] = []
    for name, col_idx in PROFILE_COLUMNS.items():
        cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
        if cell is None:
            continue
        if str(cell).strip().upper() == "X":
            present.append(name)
    return present


def _row_to_dict(row: list[Any], columns: dict[str, int]) -> dict[str, Any]:
    """Read named columns out of a row, normalising values."""
    out: dict[str, Any] = {}
    for key, col_idx in columns.items():
        cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
        out[key] = _norm(cell)
    return out


def _occurrence(sheet_tag: str, row_number: int, row: list[Any]) -> dict[str, Any]:
    occ: dict[str, Any] = {"sheet": sheet_tag, "row": row_number}
    for key in OCCURRENCE_COLUMNS:
        col_idx = EN_COLUMNS[key]
        cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
        occ[key] = _norm(cell)
    return occ


def _merge(entry: dict[str, Any], updates: dict[str, Any]) -> None:
    """Set keys on ``entry`` only when they're missing or ``None`` —
    the first non-empty value wins (we walk sheets MINIMUM → EXTENDED,
    so lower-profile entries seed the scalars and the EXTENDED sheet
    only fills gaps)."""
    for key, value in updates.items():
        if value is None:
            continue
        if entry.get(key) in (None, ""):
            entry[key] = value


def extract(xlsx: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)

    terms: dict[str, dict[str, Any]] = {}

    for sheet_name in PROFILE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(  # noqa: T201
                f"warning: sheet {sheet_name!r} not found, skipping", file=sys.stderr
            )
            continue
        sheet_tag = SHEET_TAG[sheet_name]
        ws = wb[sheet_name]
        for row_number, raw_row in enumerate(
            ws.iter_rows(min_row=5, values_only=True), start=5
        ):
            row = list(raw_row)
            bt_id_raw = row[EN_COLUMNS["id"] - 1] if row else None
            if not bt_id_raw:
                continue
            bt_id = str(bt_id_raw).strip()
            if not ID_RE.match(bt_id):
                continue

            scalars = _row_to_dict(row, EN_COLUMNS)
            scalars.update(_row_to_dict(row, FR_COLUMNS))
            scalars["profiles"] = _profiles_present(row)

            entry = terms.setdefault(bt_id, {"id": bt_id})
            _merge(entry, scalars)

            # Always record the (sheet, row) occurrence so consumers
            # can recover the per-context cardinality / xpath even when
            # the same id appears multiple times within one sheet.
            entry.setdefault("occurrences", []).append(
                _occurrence(sheet_tag, row_number, row)
            )
            sheets_present = entry.setdefault("present_in_sheets", [])
            if sheet_tag not in sheets_present:
                sheets_present.append(sheet_tag)

    return terms


def _split_xpath(xpath: str) -> list[str]:
    """Split an xpath into single-step segments, each carrying its
    leading separator.

    ``/rsm:CrossIndustryInvoice/ram:ID/@schemeID`` →
    ``["/rsm:CrossIndustryInvoice", "/ram:ID", "/@schemeID"]``.

    Element steps are separated by ``/``; attribute steps by ``/@``.
    Both are normalised so the segment string starts with its
    separator — keeping the keys self-describing when looked at out
    of context.
    """
    if not xpath.startswith("/"):
        return []
    segments: list[str] = []
    cursor = 0
    length = len(xpath)
    while cursor < length:
        # Look ahead one character past the current ``/`` to decide
        # element vs attribute step.
        is_attr = cursor + 1 < length and xpath[cursor + 1] == "@"
        # Find the start of the next step.
        next_slash = xpath.find("/", cursor + 1)
        if next_slash == -1:
            step = xpath[cursor:]
            cursor = length
        else:
            step = xpath[cursor:next_slash]
            cursor = next_slash
        if is_attr:
            # Already includes the ``/@`` prefix.
            segments.append(step)
        else:
            segments.append(step)
    return segments


def _build_tree(terms: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Re-shape the flat term registry as a tree keyed by xpath segment.

    Each tree node has the shape::

        {
            "id": "BT-31" | null,
            "name": "Seller VAT identifier" | null,
            "profiles": ["MINIMUM", ...],
            "children": {"/ram:ID": {...}, ...}
        }

    ``id`` is only populated for nodes that map exactly to a BT/BG
    occurrence — pure structural containers (the ``rsm:`` /  ``ram:``
    wrappers that the xpath traverses but which don't have their own
    BT id) leave it ``null``. When the same id appears at multiple
    xpaths (e.g. BT-2 occurs once per profile sheet) the tree records
    the union of profiles on every node it touches.
    """
    root: dict[str, Any] = {}
    for entry in terms.values():
        for occ in entry.get("occurrences", []):
            xpath = occ.get("xpath")
            if not xpath:
                continue
            segments = _split_xpath(xpath)
            if not segments:
                continue
            cursor = root
            for i, step in enumerate(segments):
                node = cursor.setdefault(
                    step, {"id": None, "name": None, "profiles": [], "children": {}}
                )
                is_leaf = i == len(segments) - 1
                if is_leaf:
                    # Attach the BT id and name to the exact xpath leaf.
                    node["id"] = entry["id"]
                    node["name"] = entry.get("name")
                # Track the profile sheets we saw this node in — both
                # for leaves and the structural parents above them.
                sheet = occ.get("sheet")
                if sheet and sheet not in node["profiles"]:
                    node["profiles"].append(sheet)
                cursor = node["children"]
    return root


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="extract_business_terms", description=__doc__.split("\n\n", maxsplit=1)[0]
    )
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=DEFAULT_XLSX,
        type=Path,
        help=(
            "Path to the Factur-X workbook "
            "(``1_FACTUR-X 1.08 - ... - VF.xlsx``). "
            f"Defaults to {DEFAULT_XLSX.relative_to(ROOT)} when run "
            "from the repository root."
        ),
    )
    parser.add_argument(
        "-o",
        "--out",
        default=DEFAULT_OUT,
        type=Path,
        help=(f"Output JSON path. Defaults to {DEFAULT_OUT.relative_to(ROOT)}."),
    )
    parser.add_argument(
        "--out-tree",
        default=DEFAULT_OUT_TREE,
        type=Path,
        help=(
            "Output JSON path for the xpath-hierarchy tree. "
            f"Defaults to {DEFAULT_OUT_TREE.relative_to(ROOT)}."
        ),
    )
    return parser


def _relative(path: Path) -> Path:
    if path.is_absolute() and path.is_relative_to(ROOT):
        return path.relative_to(ROOT)
    return path


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    xlsx: Path = args.xlsx
    out: Path = args.out
    out_tree: Path = args.out_tree

    if not xlsx.is_file():
        print(f"error: cannot read XLSX at {xlsx}", file=sys.stderr)  # noqa: T201
        return 2

    terms = extract(xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(terms, indent=2, sort_keys=True) + "\n")
    occurrences = sum(len(t["occurrences"]) for t in terms.values())
    print(  # noqa: T201
        f"wrote {len(terms)} terms ({occurrences} occurrences across "
        f"{len(PROFILE_SHEETS)} profile sheets) to {_relative(out)}"
    )

    tree = _build_tree(terms)
    out_tree.parent.mkdir(parents=True, exist_ok=True)
    out_tree.write_text(json.dumps(tree, indent=2, sort_keys=True) + "\n")

    def _count(node: dict[str, Any]) -> int:
        return 1 + sum(_count(c) for c in node["children"].values())

    node_count = sum(_count(n) for n in tree.values())
    print(  # noqa: T201
        f"wrote xpath tree with {node_count} nodes to {_relative(out_tree)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
